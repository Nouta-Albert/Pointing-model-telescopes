# -*- coding: utf-8 -*-
"""
Created on Thu Mar 20 15:04:09 2025

@author: Albert Einstein
"""


import numpy as np
import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from Tpoint_emulation_full import pointing_model, calculate_rms
import matplotlib.pyplot as plt
#import seaborn as sns
import os


### ==== CONFIGURATION SECTION (Edit This Only) ==== ###
#######*****************beware of angular difference in data***************************####################
# File path for input data
file_path = r"C:\Users\Albert Einstein\Desktop\Einstein docs\Master 2 astro\Data_analysis_thesis_report\Ecam_astrometry_pm-09-12-2024.dat"
# Output Excel file path
output_file = r"C:\Users\Albert Einstein\Desktop\Einstein docs\Master 2 astro\Data_analysis_thesis_report\comparison_results_ECAM_LM.xlsx"
# Create plots in specified directory
plot_dir = r"C:\Users\Albert Einstein\Desktop\Einstein docs\Master 2 astro\Data_analysis_thesis_report"
# Grid step in degrees
az_step = 5
el_step = 5

# Define the parameters for the two models
#params_model1 = np.array([3.999,   40.337,    90.094,   259.255,  -16.891,  -69.417,   108.328, -138.378,  14.007]) * np.pi / (180 * 3600) #For Coralie # Convert to radians
#params_model1 = np.array([5.266,   43.083,    95.647,   242.83,  61.444,  -117.706,   79.882, -91.098,  -2.955]) * np.pi / (180 * 3600) #For Ecam# Convert to radians
#params_model1 = np.array([5.6874,   42.7885,    95.3620,   259.1598,  60.1191,  -59.6833,   71.5925, -107.0097,  10.4012]) * np.pi / (180 * 3600) #For Ecam LPM # Convert to radians
#params_model1 = np.array([4.6527,   -23.3071,    106.7256,   2166.4000,  -26.2989,  5220.0105,   -1425.7723, -454.5135,  995.9181]) * np.pi / (180 * 3600)
#params_model2 = np.array([3.8827, 40.0012,   90.6151, 318.4789, -16.9759, 93.5727, 64.3222, -152.0550, 44.5881]) * np.pi / (180 * 3600) #My terms for Coralie # Convert to radians
#params_model2 = np.array([5.3268, 42.9147,   95.6802, 291.2327,  61.5283, 13.0052, 81.6729, -133.4021, 29.4735]) * np.pi / (180 * 3600) #My terms for Ecam # Convert to radians
#params_model2 = np.array([5.6639, 42.8295,   95.3157, 249.4903,  60.2507, -88.1040, 83.3708, -108.7993, 4.9913]) * np.pi / (180 * 3600) #My terms for Ecam LPM # Convert to radians
#selected_terms =         ["AN",    "AW",      "IA",     "IE",      "CA",     "TF",     "PZZ3",   "PZZ5",   "HZCZ4"]

# Model 1 parameters and selected terms
params_model1 = np.array([5.663, 43.155, 95.820, 231.245, 61.551, -106.042, 102.669, -72.235, -10.513]) * np.pi / (180 * 3600)
#params_model1 = np.array([3.999, 40.337, 90.094, 259.255, -16.891, -69.417, 108.328, -138.378, 14.007]) * np.pi / (180 * 3600) #mine coralie
#params_model1 = np.array([5.255, 42.300, 99.465, 251.014, 16.129, -56.779, 115.804, -100.679, 6.431]) * np.pi / (180 * 3600) #mine coralie
#params_model1 = np.array([5.8607, 42.8800, 95.3703, 4.8034, 60.2170, -145.2998, 7.5497, 236.9175]) * np.pi / (180 * 3600) #mine Ecam
#params_model1 = np.array([5.8607, 42.8800, 95.3703, 4.8034, 60.2170, -145.2998, 7.5497, 236.9175]) * np.pi / (180 * 3600) #mine Ecam


selected_terms_model1 = ["AN", "AW", "IA", "IE", "CA", "TF", "PZZ3", "PZZ5", "HZCZ4"]
#selected_terms_model1 = ["AN", "AW", "IA", "IE", "CA", "TX", "HZSZ2", "ECES"]

# Model 2 parameters and selected terms
#params_model2 = np.array([5.7431, 42.9427, 95.8512, 290.5862, 61.6257, 54.4379, 104.6529, -123.6786, 29.2660]) * np.pi / (180 * 3600)
#params_model2 = np.array([5.2958, 42.2573, 99.3631, 310.1095, 16.2876, 107.5932, 62.8428, -107.0813, 36.3407]) * np.pi / (180 * 3600)  #my terms coalie legacy method
#params_model2 = np.array([4.0480, 40.1406, 90.1948, 323.8330, -16.8516, 109.5270, 55.1716, -149.2950, 47.1216]) * np.pi / (180 * 3600) #tpoint coralie
#params_model2 = np.array([5.504, 42.486, 95.349, 134.131, 60.119, -112.344, 20.335, 119.374]) * np.pi / (180 * 3600) #tpoint Ecam
#params_model2 = np.array([5.682, 42.823, 95.355, 127.864, 60.125, -16.844, -11.137, -89.514, 125.689, -38.428, 9.539]) * np.pi / (180 * 3600)
params_model2 = np.array([5.6418, 43.0963, 95.7849, -50.3736, 61.6639, -114.5910, 4.1805, 293.9181]) * np.pi / (180 * 3600)



#selected_terms_model2 = ["AN", "AW", "IA", "IE", "CA", "TF", "PZZ3", "PZZ5", "HZCZ4"]
selected_terms_model2 = ["AN", "AW", "IA", "IE", "CA", "TX", "HZSZ2", "ECES"]

### ==== END CONFIGURATION ==== ###


def load_data(file_path):
    data = pd.read_csv(file_path, delim_whitespace=True, header=None, skiprows=0, 
                       names=["az_real", "el_real", "az_actual", "el_actual"])
    return np.radians(data)


def create_sky_grid(az_step, el_step):
    az = np.arange(0, 360, az_step)
    el = np.arange(30, 83, el_step)
    az_grid, el_grid = np.meshgrid(az, el)
    return np.radians(az_grid), np.radians(el_grid)


def apply_corrections(az_real, el_real, params, selected_terms):
    deltaA, deltaE = pointing_model(params, az_real, el_real, selected_terms)
    az_corrected = az_real + deltaA
    el_corrected = el_real + deltaE
    return deltaA, deltaE, az_corrected, el_corrected


def calc_rms(values):
    return np.sqrt(np.mean(np.square(values)))


def compare_models(params1, params2, terms1, terms2, output_file, az_step, el_step, data):
    az_grid, el_grid = create_sky_grid(az_step, el_step)
    az_flat = az_grid.flatten()
    el_flat = el_grid.flatten()
    
    results = {
        "az_real": np.degrees(az_flat),
        "el_real": np.degrees(el_flat),
        "deltaA_model1": [], "deltaE_model1": [], "az_corrected_model1": [], "el_corrected_model1": [],
        "deltaA_model2": [], "deltaE_model2": [], "az_corrected_model2": [], "el_corrected_model2": [],
    }

    for az, el in zip(az_flat, el_flat):
        # Model 1
        dA1, dE1, az_corr1, el_corr1 = apply_corrections(az, el, params1, terms1)
        results["deltaA_model1"].append(np.degrees(dA1) * 3600)
        results["deltaE_model1"].append(np.degrees(dE1) * 3600)
        results["az_corrected_model1"].append(np.degrees(az_corr1))
        results["el_corrected_model1"].append(np.degrees(el_corr1))

        # Model 2
        dA2, dE2, az_corr2, el_corr2 = apply_corrections(az, el, params2, terms2)
        results["deltaA_model2"].append(np.degrees(dA2) * 3600)
        results["deltaE_model2"].append(np.degrees(dE2) * 3600)
        results["az_corrected_model2"].append(np.degrees(az_corr2))
        results["el_corrected_model2"].append(np.degrees(el_corr2))

    df = pd.DataFrame(results)

    # RMS computations
    rms_az1, rms_el1 = calc_rms(results["deltaA_model1"]), calc_rms(results["deltaE_model1"])
    rms_total1 = np.sqrt(rms_az1**2 + rms_el1**2)
    res_az1, res_el1 = calculate_rms(data, params1, terms1)

    rms_az2, rms_el2 = calc_rms(results["deltaA_model2"]), calc_rms(results["deltaE_model2"])
    rms_total2 = np.sqrt(rms_az2**2 + rms_el2**2)
    res_az2, res_el2 = calculate_rms(data, params2, terms2)

    # Display
    def print_model_rms(title, az, el, total, res_az, res_el):
        print(f"\n{title} RMS Values:")
        print(f"  Azimuth RMS: {az:.2f} arcsec")
        print(f"  Elevation RMS: {el:.2f} arcsec")
        print(f"  Total RMS: {total:.2f} arcsec")
        print(f"    Residuals RMS - Azimuth: {res_az:.2f} arcsec")
        print(f"    Residuals RMS - Elevation: {res_el:.2f} arcsec")
        print(f"    Residuals RMS - total: {np.sqrt(res_el**2 + res_az**2):.2f} arcsec")

    print_model_rms("Model 1", rms_az1, rms_el1, rms_total1, res_az1, res_el1)
    print_model_rms("Model 2", rms_az2, rms_el2, rms_total2, res_az2, res_el2)

    # Save to Excel
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
        ws = writer.sheets["Results"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
            for cell in row:
                cell.number_format = "0.000"
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

    print(f"\nResults saved to {output_file}")
    return df

def visualize_key_differences(results_df, save_dir="."):
    """
    Create two publication-quality plots:
    1. Polar plot showing differences (zenith at center)
    2. 2D scatter plot showing differences
    
    Parameters:
    - results_df: DataFrame containing comparison results
    - save_dir: Directory to save plots (default current directory)
    """
    # Calculate differences
    results_df['deltaA_diff'] = results_df['deltaA_model2'] - results_df['deltaA_model1']
    results_df['deltaE_diff'] = results_df['deltaE_model2'] - results_df['deltaE_model1']
    results_df['total_diff'] = np.sqrt(results_df['deltaA_diff']**2 + results_df['deltaE_diff']**2)
    
    # Create colormap suitable for astronomical data
    from matplotlib.colors import LinearSegmentedColormap
    astro_cmap = LinearSegmentedColormap.from_list('astro', ['#2d0e4a', '#4b1d73', '#7a3b8a', '#b14e8c', '#e1717d', '#ffa47e'])
    
    # --- Plot 1: Polar Plot (Zenith at center) ---
    plt.figure(figsize=(10, 8))
    ax = plt.subplot(111, projection='polar')
    
    # Convert elevation to polar coordinates (90-el since zenith should be center)
    r = 90 - results_df['el_real']  # 0° at horizon, 90° at zenith
    theta = np.radians(results_df['az_real'])
    
    # Create polar scatter plot
    sc = ax.scatter(theta, r, c=results_df['total_diff'], 
                   cmap="rainbow", alpha=0.8, s=30, 
                   vmin=0, vmax=results_df['total_diff'].max())
    
    # Customize polar plot
    ax.set_theta_zero_location('N')  # 0° at top (North)
    ax.set_theta_direction(-1)       # Clockwise azimuth
    ax.set_rlim(0, 90)               # Zenith (0) to horizon (90)
    ax.set_rticks([0, 30, 60, 90])  # Major elevation ticks
    ax.set_yticklabels(['90°', '60°', '30°', '0°'])  # Label as elevation
    
    # Add colorbar and title
    cbar = plt.colorbar(sc, pad=0.1)
    cbar.set_label('Pointing Difference (arcsec)', rotation=270, labelpad=15)
    plt.title('Model Differences (Zenith at Center)\nAzimuth: 0°=North, 90°=East', pad=20)
    
    # Save polar plot
    polar_path = os.path.join(save_dir, "polar_differences.png")
    plt.savefig(polar_path, dpi=300, bbox_inches='tight')
    print(f"Saved polar plot to {polar_path}")
    plt.close()
    
    # --- Plot 2: 2D Scatter Plot ---
    plt.figure(figsize=(12, 6))
    
    # Create azimuth-elevation grid
    ax = plt.subplot(111)
    
    # Use hexbin for better visualization of dense points
    hb = ax.hexbin(results_df['az_real'], results_df['el_real'], 
                  C=results_df['total_diff'], gridsize=50, 
                  cmap="rainbow", reduce_C_function=np.mean)
    
    # Customize plot
    ax.set_xlabel('Azimuth (degrees)', fontsize=12)
    ax.set_ylabel('Elevation (degrees)', fontsize=12)
    ax.grid(True, linestyle='--', alpha=0.6)
    
    # Add colorbar
    cbar = plt.colorbar(hb)
    cbar.set_label('Pointing Difference (arcsec)', rotation=270, labelpad=20)
    
    # Title and layout
    plt.title('Pointing Model Differences (Model 2 - Model 1)', pad=15)
    plt.tight_layout()
    
    # Save 2D plot
    scatter_path = os.path.join(save_dir, "2d_differences.png")
    plt.savefig(scatter_path, dpi=300, bbox_inches='tight')
    print(f"Saved 2D plot to {scatter_path}")
    plt.close()

# ==== MAIN SCRIPT ====
if __name__ == "__main__":
    data = load_data(file_path)
    results_df = compare_models(params_model1, params_model2,
                              selected_terms_model1, selected_terms_model2,
                              output_file, az_step, el_step, data)
    print(results_df.head())
    visualize_key_differences(results_df, save_dir=plot_dir)


